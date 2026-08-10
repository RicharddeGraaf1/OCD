# -*- coding: utf-8 -*-
"""Genereert curatie/lijsten.xlsx — drie kale lijsten, verder niets.

  blad 1  Categorieen      naam · artikelen · aantal subcategorieen
  blad 2  Subcategorieen   naam · categorie · artikelen
  blad 3  typeBepalingen   naam · artikelen

Leest rechtstreeks uit `v2a.artikel_indeling`, dus dit is wat er wérkelijk in
de database staat en niet een herberekening die ernaast kan gaan lopen. Draai
eerst `bouw_indeling.py`.

De indeling zelf komt uit `curatie/subcategorie_regels.py`; dáár pas je iets
aan, niet hier.

Draaien:  python scripts/genereer_lijsten.py
"""
import os

import psycopg
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Doelwit-DB in dezelfde volgorde als de rest van de sync: `full_sync.py`
# zet OCD_DB_URL wanneer je met --target prod rechtstreeks tegen productie
# draait. Wie hier alleen DSO_DB leest, herbouwt bij zo'n run stilzwijgend de
# LOKALE indeling en laat prod onaangeroerd.
DB = (os.environ.get("OCD_DB_URL")
      or os.environ.get("DSO_DB")
      or "postgresql://postgres:postgres@localhost:5434/dso")
UIT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..",
                   "curatie", "lijsten.xlsx")


def main():
    conn = psycopg.connect(DB)
    cur = conn.cursor()

    cur.execute("""SELECT categorie, count(*), count(DISTINCT subcategorie)
                   FROM v2a.artikel_indeling WHERE categorie IS NOT NULL
                   GROUP BY 1 ORDER BY 2 DESC""")
    cats = cur.fetchall()

    cur.execute("""SELECT subcategorie, min(categorie), count(*)
                   FROM v2a.artikel_indeling WHERE subcategorie IS NOT NULL
                   GROUP BY 1 ORDER BY 3 DESC""")
    subs = cur.fetchall()

    cur.execute("""SELECT type_bepaling, count(*) FROM v2a.artikel_indeling
                   WHERE type_bepaling IS NOT NULL GROUP BY 1 ORDER BY 2 DESC""")
    types = cur.fetchall()

    cur.execute("""SELECT count(*), count(*) FILTER (WHERE categorie IS NULL),
                          count(*) FILTER (WHERE type_bepaling IS NULL)
                   FROM v2a.artikel_indeling""")
    totaal, geen_cat, geen_type = cur.fetchone()

    wb = Workbook()
    wb.remove(wb.active)
    KOP, VUL = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="44546A")

    def blad(naam, kolommen, rijen):
        ws = wb.create_sheet(naam)
        ws.append([k for k, _ in kolommen])
        for i, (_, br) in enumerate(kolommen, 1):
            ws.cell(row=1, column=i).font = KOP
            ws.cell(row=1, column=i).fill = VUL
            ws.column_dimensions[get_column_letter(i)].width = br
        for r in rijen:
            ws.append(list(r))
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    blad("Categorieen", [("categorie", 26), ("artikelen", 12), ("subcategorieen", 15)],
         cats + [("(niet ingedeeld)", geen_cat, "")])
    blad("Subcategorieen", [("subcategorie", 38), ("categorie", 24), ("artikelen", 12)], subs)
    blad("typeBepalingen", [("typeBepaling", 30), ("artikelen", 12)],
         types + [("(niet herkend)", geen_type)])

    pad = os.path.abspath(UIT)
    os.makedirs(os.path.dirname(pad), exist_ok=True)
    try:
        wb.save(pad)
        print("geschreven:", pad)
    except PermissionError:
        wijk = pad.replace(".xlsx", "-nieuw.xlsx")
        wb.save(wijk)
        print(f"LET OP: {pad} staat open; weggeschreven als {wijk}")

    print(f"\n{totaal} artikelen · {len(cats)} categorieen · "
          f"{len(subs)} subcategorieen · {len(types)} typeBepalingen")
    print(f"  ingedeeld    : {totaal-geen_cat} ({100*(totaal-geen_cat)/totaal:.1f}%)")
    print(f"  typeBepaling : {totaal-geen_type} ({100*(totaal-geen_type)/totaal:.1f}%)")
    print("\nSUBCATEGORIEEN")
    for s, c, n in subs:
        print(f"  {n:7}  {s[:36]:38} {c}")
    conn.close()


if __name__ == "__main__":
    main()
