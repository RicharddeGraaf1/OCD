# -*- coding: utf-8 -*-
"""Bouwt v2a.pad_categorie + v2a.artikel_indeling — de indeling als opzoeking.

Twee assen, allebei een opzoeking op genormaliseerde tekst:

  categorie / subcategorie  <- VOOROUDERPAD        <- handcuratie (xlsx)
  type_bepaling             <- ARTIKELOPSCHRIFT    <- gesloten lexicon hieronder

Geen embeddings, geen clustering, geen centroïdes. Het opschriftpad is een
bruikbare joinsleutel: 10.798 ruwe paden vallen na normaliseren samen tot
9.999, en de 1.000 grootste dekken 78,6% van alle artikelen.

NULL = niet ingedeeld, en dat is een geldig antwoord (gebruiker 2026-08-09).
Geen sleuteltreffer betekent geen waarde — niet raden, niet naar de
dichtstbijzijnde buur schuiven.

De typeBepaling-as heeft GEEN curatie nodig en kan dus vooruitlopen op het
invullen van de xlsx.

Twee doelen, dezelfde regels:

  v2a.artikel_indeling    de vigerende regelingen (p2p)
  v2a.wijziging_indeling  de renvooi-tekst van ontwerpen en besluiten
                          (p2pwijziging), voor de wijzigingentour

Die tweede fase hoort hier en niet in een eigen script: liepen de regelsets
uiteen, dan zou hetzelfde artikel in het register een ander onderwerp krijgen
dan in de tour ernaast.

Draaien:
    python scripts/bouw_indeling.py                     # alles
    python scripts/bouw_indeling.py --alleen-type       # zonder de xlsx
    python scripts/bouw_indeling.py --alleen-wijzigingen
    python scripts/bouw_indeling.py --ddl               # tabellen (her)aanmaken
"""
import argparse
import os
import re
import sys
import unicodedata
from collections import Counter

import psycopg

# Doelwit-DB in dezelfde volgorde als de rest van de sync: `full_sync.py`
# zet OCD_DB_URL wanneer je met --target prod rechtstreeks tegen productie
# draait. Wie hier alleen DSO_DB leest, herbouwt bij zo'n run stilzwijgend de
# LOKALE indeling en laat prod onaangeroerd.
DB = (os.environ.get("OCD_DB_URL")
      or os.environ.get("DSO_DB")
      or "postgresql://postgres:postgres@localhost:5434/dso")
HIER = os.path.dirname(os.path.abspath(__file__))
DDL_PAD = os.path.join(HIER, "2026-08-add-pad-categorie.sql")
DDL_WIJZ = os.path.join(HIER, "2026-08-add-wijziging-indeling.sql")
XLSX = os.path.join(HIER, "..", "curatie", "lijsten.xlsx")
CURATIE_VERSIE = "v2-2026-08"

# ── De gesloten typeBepaling-lijst ───────────────────────────────────────────
# Volgorde telt: de eerste treffer wint, dus specifieke patronen boven
# generieke. Afgeleid uit het lexicon van de fase-0-dekkingsmeting, dat op
# 95,4% van 879.348 eenheden een uitspraak deed.
#
# Links het patroon dat op het GENORMALISEERDE artikelopschrift wordt gematcht,
# rechts de canonieke waarde. Meerdere patronen mogen naar dezelfde waarde
# wijzen — dat is precies het punt van een gesloten lijst.
TYPE_BEPALING = [
    (r"toepassingsbereik|toepasselijkheid|reikwijdte|werkingssfeer", "toepassingsbereik"),
    (r"oogmerk(en)?|doel(einden)?( van de regels)?", "oogmerk"),
    (r"begripsbepaling(en)?|begrippen|definities", "begripsbepaling"),
    (r"normadressaat", "normadressaat"),
    (r"(specifieke )?zorgplicht", "zorgplicht"),
    (r"maatwerk(voorschrift(en)?|regels)?", "maatwerkvoorschrift"),
    (r"gelijkwaardigheid", "gelijkwaardigheid"),
    (r"meet-? ?en ?rekenbepalingen|meetmethoden?|metingen?|meten", "meet- en rekenbepaling"),
    (r"gegevens en bescheiden|informatieplicht|informeren", "gegevens en bescheiden"),
    (r"meldingsplicht|melding(en)?", "meldingsplicht"),
    (r"aanvraagvereisten", "aanvraagvereisten"),
    (r"(aanwijzing )?vergunningplicht(ige gevallen)?", "vergunningplicht"),
    (r"beoordelingsregels?.*", "beoordelingsregel"),
    (r"overgangsrecht.*|eerbiedigende werking.*", "overgangsrecht"),
    (r"voorrangsbepaling", "voorrangsbepaling"),
    (r"slotbepaling(en)?|slotregel|citeertitel|inwerkingtreding", "slotbepaling"),
    (r"gereserveerd|vervallen", "gereserveerd of vervallen"),
    (r"verbod(sbepaling)?|gebod", "verbod of gebod"),
    (r"strafbepaling|handhaving|toezicht", "handhaving"),
    (r"bevoegd gezag", "bevoegd gezag"),
    # Wro-redactie: dit zijn typeBepalingen, het onderwerp is de bestemmingsnaam
    (r"bestemmingsomschrijving", "bestemmingsomschrijving"),
    (r"afwijken van de (bouw|gebruiks)regels|afwijkingsregels", "afwijkingsregel"),
    (r"specifieke gebruiksregels|gebruiksregels", "gebruiksregels"),
    (r"bouwregels", "bouwregels"),
    (r"wijzigingsbevoegdheid", "wijzigingsbevoegdheid"),
    (r"nadere eisen", "nadere eisen"),
    (r"aanlegregels", "aanlegregel"),
    (r"wijze van meten", "wijze van meten"),
    (r"anti-?dubbeltelregel", "anti-dubbeltelregel"),
]
TYPE_RE = [(re.compile(r"^(?:" + p + r")$"), w) for p, w in TYPE_BEPALING]


def kern(seg: str) -> str:
    """Artikelopschrift -> matchbare kern.

    Nummering eraf, en bij een dubbele punt de LINKERkant houden voor het
    onderwerp maar de RECHTERkant voor het type — vandaar dat deze functie
    beide teruggeeft via kern_links/kern_rechts hieronder.
    """
    s = unicodedata.normalize("NFKD", seg or "").encode("ascii", "ignore").decode()
    s = re.sub(r"\s+", " ", s.strip().lower())
    s = re.sub(r"^(?:artikel\s+|hoofdstuk\s+|paragraaf\s+)?[0-9]+(?:[.\-][0-9a-z]+)*\.?\s*", "", s)
    return s.strip(" .:-()[]")


def bepaal_type(opschrift: str):
    """Canonieke typeBepaling, of None.

    Probeert eerst het hele opschrift, dan de kant NA een dubbele punt
    ("Geur landbouwhuisdieren: eerbiedigende werking" -> overgangsrecht), dan
    de kant ervoor ("Toepassingsbereik: eerbiedigende werking").
    """
    k = kern(opschrift)
    if not k:
        return None
    kandidaten = [k]
    if ":" in k:
        links, rechts = k.split(":", 1)
        kandidaten += [rechts.strip(), links.strip()]
    for kand in kandidaten:
        for rx, waarde in TYPE_RE:
            if rx.match(kand):
                return waarde
    return None


def pad_sleutel(pad: str) -> str:
    """Ruw voorouderpad -> stabiele joinsleutel.

    Klapt 10.798 ruwe paden samen tot 9.999: accenten, hoofdletters,
    leestekens en nummering verschillen per bronhouder, de rest niet.
    """
    s = unicodedata.normalize("NFKD", pad or "").encode("ascii", "ignore").decode()
    s = s.lower().strip()
    s = re.sub(r"[^a-z0-9> ]", "", s)
    s = re.sub(r"\b\d+(\.\d+)*\b", "", s)
    return re.sub(r"\s+", " ", s).strip()




# ── Gedeeld met genereer_lijsten.py ────────────────────────────────────

BEPALINGTYPE = [
    r"toepassingsbereik", r"toepasselijkheid", r"reikwijdte", r"werkingssfeer",
    r"oogmerk(en)?", r"doel(einden)?( van de regels)?",
    r"begripsbepaling(en)?", r"begrippen", r"definities",
    r"normadressaat", r"(specifieke )?zorgplicht",
    r"maatwerk(voorschrift(en)?|regels)?", r"gelijkwaardigheid",
    r"meet-? ?en ?rekenbepalingen", r"meetmethoden?", r"metingen?", r"meten",
    r"gegevens en bescheiden", r"informatieplicht", r"informeren",
    r"melding(splicht|en)?", r"aanvraagvereisten",
    r"(aanwijzing )?vergunningplicht(ige gevallen)?",
    r"verbod(sbepaling)?", r"gebod",
    r"algemene (bepalingen|regels|bouwregels|gebruiksregels|afwijkingsregels|procedureregels|aanduidingsregels)",
    r"overgangsrecht.*", r"eerbiedigende werking.*", r"voorrangsbepaling",
    r"slotbepaling(en)?", r"slotregel", r"citeertitel", r"inwerkingtreding",
    r"gereserveerd", r"vervallen", r"procedure(regels)?", r"beoordelingsregels?.*",
    r"bevoegd gezag", r"strafbepaling", r"handhaving", r"toezicht",
    r"bestemmingsomschrijving", r"bouwregels", r"(specifieke )?gebruiksregels",
    r"afwijken van de (bouw|gebruiks)regels", r"afwijkingsregels",
    r"wijzigingsbevoegdheid", r"nadere eisen", r"aanlegregels",
    r"wijze van meten", r"anti-?dubbeltelregel", r"inleidende regels",
    r"bestemmingsregels", r"overgangs-? ?en ?slotregels",
    r"artikelsgewijze toelichting", r"algemene toelichting", r"toelichting",
    r"bijlage[n]?( \w+)?", r"inleiding", r"samenvatting", r"leeswijzer",
    r"inhoudsopgave", r"overige bepalingen", r"overig(e)?", r"algemeen",
]


BEPALINGTYPE_RE = re.compile(r"^(?:" + "|".join(BEPALINGTYPE) + r")$")


# Containers: wél een onderwerp, maar te grof om als label te dienen.
GENERIEK = {"activiteiten", "overige activiteiten", "thematische activiteiten",
            "regels", "planregels", "voorschriften", "bepalingen",
            "activiteitenbruidsschat"}

# Structuur, geen onderwerp. Krijgen een expliciet voorstel om NIET in te delen —
# beter een lege categorie dan een knop "bruidsschat" naast "geur".
STRUCTUUR = {"bruidsschat", "activiteiten bruidsschat", "thema's", "themas",
             "voormalige rijksregels", "inhoudelijke regels",
             "voorbeschermingsregels", "tijdelijk deel", "hoofdstukindeling"}



def is_bepalingtype(seg):
    k = kern(seg)
    return len(k) < 4 or bool(BEPALINGTYPE_RE.match(k))


def bruikbaar(seg):
    k = kern(seg)
    return bool(k) and not is_bepalingtype(seg) and k not in GENERIEK




def label_van(pad_segs):
    """Voorstel-label = het KORTSTE bruikbare kopje in de keten.

    Niet het diepste: dat is de volzin waar dit script eerder op strandde.
    Niet het ondiepste: dat is de boekdeel-container. Het kortste is in de
    praktijk het kopje dat de wetgever als onderwerpsnaam bedoelde — 'Geur',
    'Bodembeheer', 'Trillingen'.
    """
    kand = [kern(s) for s in pad_segs if bruikbaar(s)]
    return min(kand, key=len) if kand else ""




# Schrijfwijzen van hetzelfde IMOW-thema. `Water`, `water en watersystemen` en
# `WaterEnWatersystemen` komen alle drie voor in p2p.juridische_regel.thema.
SAMEN = {"water en watersystemen": "water", "waterenwatersystemen": "water",
         "bouwwerken": "bouwen", "cultureel erfgoed": "erfgoed",
         "cultureelerfgoed": "erfgoed", "milieualgemeen": "milieu",
         "landgebruik": "planologisch gebruik",
         "energieennatuurlijkehulpbronnen": "energie"}


def norm_thema(t):
    s = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", t or "").lower().strip()
    return SAMEN.get(s, SAMEN.get(s.replace(" ", ""), s))




THEMA_SQL = """
WITH RECURSIVE keten AS (
    SELECT a.id AS art, a.regeling_expression, a.parent_id, ARRAY[]::text[] AS ops, 0 AS d
    FROM p2p.tekst_element a WHERE a.element_type = 'Artikel'
  UNION ALL
    SELECT k.art, k.regeling_expression, p.parent_id,
           coalesce(p.opschrift,'') || k.ops, k.d + 1
    FROM keten k JOIN p2p.tekst_element p ON p.id = k.parent_id WHERE k.d < 8
), plat AS (
    SELECT DISTINCT ON (art) art, regeling_expression,
           array_to_string(array_remove(ops,''), ' > ') AS pad
    FROM keten ORDER BY art, d DESC
)
SELECT pl.pad, unnest(jr.thema) AS thema, count(*) AS n
FROM plat pl
JOIN p2p.regeling r ON r.frbr_expression = pl.regeling_expression
JOIN p2p.tekst_element lid ON lid.parent_id = pl.art
JOIN p2p.juridische_regel jr ON jr.regeltekst_wid = lid.wid
     AND jr.regeling_expression = lid.regeling_expression
WHERE NOT coalesce(r.inactief, false)
  AND jr.thema IS NOT NULL AND array_length(jr.thema, 1) > 0
GROUP BY 1, 2
"""


PAD_SQL = """
WITH RECURSIVE keten AS (
    SELECT a.id AS art, a.regeling_expression, a.parent_id, a.wid, a.opschrift AS art_op,
           ARRAY[]::text[] AS ops, 0 AS d
    FROM p2p.tekst_element a WHERE a.element_type = 'Artikel'
  UNION ALL
    SELECT k.art, k.regeling_expression, p.parent_id, k.wid, k.art_op,
           coalesce(p.opschrift,'') || k.ops, k.d + 1
    FROM keten k JOIN p2p.tekst_element p ON p.id = k.parent_id WHERE k.d < 8
)
SELECT k.art, k.regeling_expression, k.wid, k.art_op,
       array_to_string(array_remove(k.ops,''), ' > ') AS pad,
       r.bronhouder
FROM (SELECT DISTINCT ON (art) art, regeling_expression, wid, art_op, ops
      FROM keten ORDER BY art, d DESC) k
JOIN p2p.regeling r ON r.frbr_expression = k.regeling_expression
WHERE NOT coalesce(r.inactief, false)
"""


# Zelfde klim als PAD_SQL, maar door de renvooi-boom. Scope = artikelen die in
# dit besluit iets doen (wijzigactie / vervallen / renvooi), want alleen die
# komen in de tour. De besluit-join levert regeling_work: p2pwijziging is per
# ontwerpbesluit gevuld en heeft geen expression van de geldende regeling.
WIJZIGING_PAD_SQL = """
WITH RECURSIVE elems AS (
    SELECT t.id, t.parent_id, t.wid, t.opschrift, t.element_type,
           b.regeling_work, t.wijzigactie, t.vervallen, t.bevat_renvooi
    FROM p2pwijziging.tekst_element t
    JOIN p2pwijziging.besluit b USING (ontwerpbesluit_id)
),
keten AS (
    SELECT e.id AS art, e.regeling_work, e.parent_id, e.wid, e.opschrift AS art_op,
           ARRAY[]::text[] AS ops, 0 AS d
    FROM elems e
    WHERE e.element_type = 'Artikel'
      AND (e.wijzigactie IS NOT NULL OR e.vervallen OR e.bevat_renvooi)
  UNION ALL
    SELECT k.art, k.regeling_work, p.parent_id, k.wid, k.art_op,
           coalesce(p.opschrift,'') || k.ops, k.d + 1
    FROM keten k JOIN elems p ON p.id = k.parent_id WHERE k.d < 8
)
SELECT DISTINCT ON (art) art, regeling_work, wid, art_op,
       array_to_string(array_remove(ops,''), ' > ') AS pad
FROM keten ORDER BY art, d DESC
"""

# De vigerende indeling als route 1. `regeling_expression` is een expression,
# p2pwijziging kent alleen de work — vandaar het afkappen op '/nld@'.
REGISTER_SQL = """
SELECT split_part(regeling_expression, '/nld@', 1) AS work, wid,
       categorie, subcategorie
FROM v2a.artikel_indeling
WHERE categorie IS NOT NULL AND wid IS NOT NULL
"""


def _laad_regels():
    """De curatie uit curatie/subcategorie_regels.py.

    Losse module en geen constante hier, zodat een oordeel over de taxonomie
    aanpasbaar is zonder dit script te openen — en zodat in git zichtbaar is
    wanneer de indeling wijzigde en wanneer de machinerie.
    """
    import importlib.util
    pad = os.path.join(HIER, "..", "curatie", "subcategorie_regels.py")
    spec = importlib.util.spec_from_file_location("subcategorie_regels", pad)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return ([re.compile(p) for p in mod.NIET_INDELEN],
            [(re.compile(p), c, s) for p, c, s in mod.REGELS],
            mod.CONTAINER)


NIET_RE, REGEL_RE, CONTAINER = _laad_regels()


def kort(seg: str) -> str:
    """Volzin -> label: knip bij het eerste scheidingsteken, cap op 38."""
    k = kern(seg)
    for sep in (", ", ": ", " anders dan", " als bedoeld", " voor zover",
                " in samenhang", " zoals bedoeld"):
        if sep in k:
            k = k.split(sep)[0].strip()
    return k[:38].strip()


def ruw_label(pad_segs):
    """Het kopje uit de keten dat het onderwerp het beste benoemt.

    Kortste bruikbare kopje dat geen container is. Zijn er alleen containers,
    dan het diepste kopje, ingekort — anders wordt een boekdeel het label.
    """
    kand = [s for s in pad_segs if bruikbaar(s)]
    schoon = [kern(s) for s in kand if kern(s) not in CONTAINER]
    if schoon:
        return min(schoon, key=len)
    return kort(kand[-1]) if kand else ""


def cureer(label: str):
    """(categorie, subcategorie), of None = niet ingedeeld."""
    if not label:
        return None
    for rx in NIET_RE:
        if rx.search(label):
            return None
    for rx, cat, sub in REGEL_RE:
        if rx.search(label):
            return (cat, sub)
    return None


def lees_xlsx(pad):
    """Gecureerde labels uit het werkblad. Ontbreekt het bestand of is het nog
    niet ingevuld, dan is dat geen fout: de typeBepaling-as draait door en de
    categorie-as blijft leeg (= niet ingedeeld)."""
    if not os.path.exists(pad):
        print(f"  {pad} bestaat niet — categorie-as blijft leeg")
        return {}
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl ontbreekt: pip install openpyxl")
    wb = load_workbook(pad, read_only=True, data_only=True)
    if "Paden" not in wb.sheetnames:
        # Normale situatie sinds 2026-08-09: de indeling komt uit
        # curatie/subcategorie_regels.py. Dit blad is de uitzonderingsroute
        # voor een pad dat je met de hand anders wilt hebben dan de regels
        # bepalen; bestaat het niet, dan is er simpelweg geen uitzondering.
        print("  geen handmatige pad-uitzonderingen (indeling komt uit de regels)")
        return {}
    ws = wb["Paden"]
    rijen = ws.iter_rows(values_only=True)
    kop = [str(c or "").strip().lower() for c in next(rijen)]
    idx = {n: kop.index(n) for n in ("bevestigd", "categorie", "subcategorie",
                                     "opschriftpad", "voorkomens", "bronhouders")
           if n in kop}
    if "opschriftpad" not in idx:
        print("  kolom 'opschriftpad' ontbreekt — categorie-as blijft leeg")
        return {}
    if "bevestigd" not in idx:
        sys.exit("  kolom 'bevestigd' ontbreekt — genereer het werkblad opnieuw.\n"
                 "  Zonder die kolom is voorinvulling niet van curatie te "
                 "onderscheiden, en dat is precies de fout die dit project "
                 "moet oplossen.")
    uit, leeg, onbevestigd = {}, 0, 0
    for r in rijen:
        ruw = r[idx["opschriftpad"]]
        if not ruw:
            continue
        # Alleen rijen die een mens heeft afgevinkt. De generator vult
        # `subcategorie` altijd voor en `categorie` soms; zonder deze eis zou
        # die voorinvulling als curatie de database in lopen.
        if not str(r[idx["bevestigd"]] or "").strip():
            onbevestigd += 1
            continue
        cat = (r[idx["categorie"]] or "").strip() if "categorie" in idx else ""
        sub = (r[idx["subcategorie"]] or "").strip() if "subcategorie" in idx else ""
        if not cat:
            leeg += 1      # afgevinkt maar bewust zonder categorie = niet ingedeeld
            continue
        uit[pad_sleutel(str(ruw))] = {
            "categorie": cat or None,
            "subcategorie": sub or None,
            "voorbeeld": str(ruw),
            "n_art": r[idx["voorkomens"]] if "voorkomens" in idx else None,
            "n_bh": r[idx["bronhouders"]] if "bronhouders" in idx else None,
        }
    print(f"  {len(uit)} bevestigde paden gelezen "
          f"({onbevestigd} nog niet afgevinkt, {leeg} afgevinkt zonder categorie)")
    return uit


def bouw_wijzigingen(cur, curatie: dict) -> None:
    """Vult v2a.wijziging_indeling — de indeling op de renvooi-tekst.

    Draait ná de artikel-indeling en leest die: route 1 is de vigerende
    indeling op (work, wid), route 2 het opschriftpad uit de renvooi-boom.
    Alleen route 2 kent nieuwe artikelen; alleen route 1 houdt de tour gelijk
    aan wat het register bij hetzelfde artikel toont.
    """
    print("\nwijzigingen indelen...", flush=True)
    # Zelf aanmaken als de tabel er nog niet is, zodat dit ook draait wanneer
    # full_sync.py het script zonder --ddl aanroept. De DDL is DROP+CREATE, dus
    # hij mag níet onvoorwaardelijk lopen.
    cur.execute("SELECT to_regclass('v2a.wijziging_indeling')")
    if cur.fetchone()[0] is None:
        cur.execute(open(DDL_WIJZ, encoding="utf-8").read())
        print("  v2a.wijziging_indeling aangemaakt")

    cur.execute(REGISTER_SQL)
    register = {}
    for work, wid, cat, sub in cur.fetchall():
        register.setdefault((work, wid), (cat, sub))
    print(f"  {len(register)} ingedeelde artikelen in het register als route 1")

    cur.execute(WIJZIGING_PAD_SQL)
    rijen, tel, gezien = [], Counter(), set()
    for _art, work, wid, art_op, pad in cur.fetchall():
        if wid is None or (work, wid) in gezien:
            continue
        gezien.add((work, wid))
        sleutel = pad_sleutel(pad)

        hit = register.get((work, wid))
        if hit:
            cat, sub, bron_cat = hit[0], hit[1], "register"
        else:
            # Zelfde volgorde als in de artikel-indeling: handmatige
            # pad-koppeling boven de regels.
            pc = curatie.get(sleutel)
            if pc:
                cat, sub, bron_cat = pc["categorie"], pc["subcategorie"], "renvooi"
            else:
                uit = cureer(ruw_label(pad.split(" > "))) if pad else None
                cat, sub = uit if uit else (None, None)
                bron_cat = "renvooi" if uit else None

        tb = bepaal_type(art_op)
        rijen.append((work, wid, sleutel, cat, sub, tb, bron_cat, CURATIE_VERSIE))
        tel["totaal"] += 1
        tel["met_categorie"] += bool(cat)
        tel["met_type"] += bool(tb)
        tel[f"via_{bron_cat}"] += bool(cat)

    cur.execute("TRUNCATE v2a.wijziging_indeling")
    with cur.copy("""COPY v2a.wijziging_indeling
        (regeling_work, artikel_wid, pad_sleutel, categorie, subcategorie,
         type_bepaling, herkomst, curatie_versie) FROM STDIN""") as cp:
        for r in rijen:
            cp.write_row(r)

    t = tel["totaal"] or 1
    print(f"  gewijzigde artikelen : {tel['totaal']}")
    print(f"  met categorie        : {tel['met_categorie']:7} ({100*tel['met_categorie']/t:5.1f}%)"
          f"  — register {tel['via_register']}, renvooi {tel['via_renvooi']}")
    print(f"  met typeBepaling     : {tel['met_type']:7} ({100*tel['met_type']/t:5.1f}%)")


def main(ddl: bool, alleen_type: bool, alleen_wijzigingen: bool) -> None:
    conn = psycopg.connect(DB, autocommit=True)
    cur = conn.cursor()

    if ddl:
        cur.execute(open(DDL_PAD, encoding="utf-8").read())
        cur.execute(open(DDL_WIJZ, encoding="utf-8").read())
        print("tabellen aangemaakt")

    curatie = {} if alleen_type else lees_xlsx(os.path.abspath(XLSX))

    if alleen_wijzigingen:
        bouw_wijzigingen(cur, curatie)
        conn.close()
        return

    if curatie:
        cur.execute("TRUNCATE v2a.pad_categorie CASCADE")
        cur.executemany(
            """INSERT INTO v2a.pad_categorie
               (pad_sleutel, pad_voorbeeld, categorie, subcategorie,
                n_artikelen, n_bronhouders, bron, curatie_versie)
               VALUES (%s,%s,%s,%s,%s,%s,'curatie',%s)""",
            [(k, v["voorbeeld"], v["categorie"], v["subcategorie"],
              v["n_art"], v["n_bh"], CURATIE_VERSIE) for k, v in curatie.items()])
        print(f"  v2a.pad_categorie gevuld: {len(curatie)} sleutels")

    print("artikelen inlezen en indelen...", flush=True)
    cur.execute(PAD_SQL)
    rijen, tel = [], Counter()
    for art, expr, wid, art_op, pad, _bron in cur.fetchall():
        sleutel = pad_sleutel(pad)
        # Eerst de handmatige koppeling op het exacte pad (die wint altijd),
        # anders de curatieregels op het label uit de keten.
        hit = curatie.get(sleutel)
        if hit:
            cat, sub, bron_cat = hit["categorie"], hit["subcategorie"], "pad-curatie"
        else:
            uit = cureer(ruw_label(pad.split(" > ")))
            cat, sub = uit if uit else (None, None)
            bron_cat = "regels" if uit else None
        tb = bepaal_type(art_op)
        herkomst = ("beide" if cat and tb else
                    bron_cat if cat else
                    "artikelopschrift" if tb else None)
        rijen.append((art, expr, wid, sleutel, cat, sub, tb, herkomst, CURATIE_VERSIE))
        tel["totaal"] += 1
        tel["met_categorie"] += bool(cat)
        tel["met_type"] += bool(tb)
        tel["niets"] += not (cat or tb)

    # pad_categorie wordt de gematerialiseerde uitkomst van de regels: één rij
    # per uniek pad, met wat eruit kwam. Niet nodig voor de werking — de
    # indeling zit al in artikel_indeling — maar wel om in SQL te kunnen zien
    # WAAROM een artikel zijn categorie kreeg, zonder het script te draaien.
    per_pad = {}
    for art, expr, wid, sleutel, cat, sub, tb, herkomst, versie in rijen:
        per_pad.setdefault(sleutel, [cat, sub, 0])
        per_pad[sleutel][2] += 1
    cur.execute("TRUNCATE v2a.pad_categorie CASCADE")
    cur.executemany(
        """INSERT INTO v2a.pad_categorie
           (pad_sleutel, categorie, subcategorie, n_artikelen, bron, curatie_versie)
           VALUES (%s,%s,%s,%s,%s,%s)""",
        [(s, c, sub, n, "curatie" if s in curatie else "regels", CURATIE_VERSIE)
         for s, (c, sub, n) in per_pad.items()])
    print(f"  v2a.pad_categorie: {len(per_pad)} paden")

    cur.execute("TRUNCATE v2a.artikel_indeling")
    with cur.copy("""COPY v2a.artikel_indeling
        (tekst_element_id, regeling_expression, wid, pad_sleutel, categorie,
         subcategorie, type_bepaling, herkomst, curatie_versie) FROM STDIN""") as cp:
        for r in rijen:
            cp.write_row(r)

    t = tel["totaal"]
    print(f"\n  artikelen ingedeeld : {t}")
    print(f"  met categorie       : {tel['met_categorie']:7} ({100*tel['met_categorie']/t:5.1f}%)")
    print(f"  met typeBepaling    : {tel['met_type']:7} ({100*tel['met_type']/t:5.1f}%)")
    print(f"  helemaal niets      : {tel['niets']:7} ({100*tel['niets']/t:5.1f}%)")

    print("\n  verdeling typeBepaling:")
    cur.execute("""SELECT coalesce(type_bepaling,'(niet herkend)'), count(*)
                   FROM v2a.artikel_indeling GROUP BY 1 ORDER BY 2 DESC""")
    for waarde, n in cur.fetchall():
        print(f"    {n:7}  {waarde}")

    bouw_wijzigingen(cur, curatie)
    conn.close()


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--ddl", action="store_true", help="tabellen (her)aanmaken")
    p.add_argument("--alleen-type", action="store_true", help="xlsx negeren")
    p.add_argument("--alleen-wijzigingen", action="store_true",
                   help="alleen v2a.wijziging_indeling herbouwen")
    a = p.parse_args()
    main(a.ddl, a.alleen_type, a.alleen_wijzigingen)
